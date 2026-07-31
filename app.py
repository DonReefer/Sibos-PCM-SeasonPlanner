"""
PCM Season Planner 2026 – editierbare Planung, graues Design,
fixierte Fahrerspalte, Eignungs- & Fahrerwerte-Farben
"""

import streamlit as st
import pandas as pd
import openpyxl
from pathlib import Path
import io
import re
from datetime import datetime
import html as html_lib

from st_aggrid import AgGrid, GridOptionsBuilder, JsCode, DataReturnMode, GridUpdateMode

st.set_page_config(
    page_title="PCM Season Planner 2026",
    page_icon="🚴",
    layout="wide",
    initial_sidebar_state="collapsed",
)

BASE = Path(__file__).parent
EXCEL = BASE / "PCM_Saisonplaner.xlsx"
LOGO = BASE / "logo.png"
MAX_RIDERS = 7
MAX_RENNTAGE = 70

st.markdown("""
<style>
  header[data-testid="stHeader"] { background: transparent !important; height: 0 !important; }
  [data-testid="stToolbar"], [data-testid="stDecoration"] { display: none !important; }
  #MainMenu, footer { visibility: hidden; }
  .stApp { background: #e8eaed !important; color: #1f2937 !important; }
  .block-container { padding-top: 1.5rem !important; padding-bottom: 2rem !important; max-width: 100% !important; }
  .stButton > button {
    background: #f3f4f6 !important; color: #1f2937 !important;
    border: 1.5px solid #9ca3af !important; border-radius: 8px !important;
    font-weight: 500 !important; height: 2.4rem !important; box-shadow: none !important;
  }
  .stButton > button:hover { background: #e5e7eb !important; border-color: #6b7280 !important; }
  .stButton > button[kind="primary"] {
    background: #4b5563 !important; border-color: #374151 !important;
    color: #f9fafb !important; font-weight: 600 !important;
  }
  .stButton > button[kind="secondary"] {
    background: #e5e7eb !important; border-color: #9ca3af !important;
    color: #374151 !important; font-size: 0.88rem !important;
  }
  .stTextInput input, .stSelectbox [data-baseweb="select"] > div {
    background: #f9fafb !important; border: 1.5px solid #9ca3af !important;
    border-radius: 8px !important; color: #1f2937 !important;
  }
  .stTabs [data-baseweb="tab-list"] {
    background: #f3f4f6; border-radius: 10px; gap: 3px; padding: 4px; border: 1.5px solid #9ca3af;
  }
  .stTabs [data-baseweb="tab"] { color: #4b5563; border-radius: 8px; font-size: 0.88rem; }
  .stTabs [aria-selected="true"] { background: #4b5563 !important; color: #f9fafb !important; }
  div[data-testid="stDataFrame"] {
    background: #f3f4f6 !important; border: 1.5px solid #9ca3af !important; border-radius: 8px !important;
  }
  .import-hint {
    background: #f3f4f6; border: 1.5px solid #9ca3af; border-radius: 10px;
    padding: 12px 14px; font-size: 0.9rem; color: #374151; margin-top: 8px;
  }
  .plan-footer { font-size: 0.8rem; color: #4b5563; padding: 6px 4px; }
  /* AgGrid hellgrau */
  .ag-theme-streamlit, .ag-theme-alpine {
    --ag-background-color: #f3f4f6;
    --ag-header-background-color: #e5e7eb;
    --ag-odd-row-background-color: #eef0f3;
    --ag-row-hover-color: #e5e7eb;
    --ag-selected-row-background-color: #d1d5db;
    --ag-border-color: #c4c9d1;
    --ag-header-foreground-color: #1f2937;
    --ag-foreground-color: #1f2937;
  }
</style>
""", unsafe_allow_html=True)


def parse_date_range(s, year=2026):
    if not s or not isinstance(s, str):
        return None, None
    s = s.strip().rstrip(".")
    m = re.match(r"^(\d{1,2})\.(\d{1,2})$", s)
    if m:
        try:
            d = datetime(year, int(m.group(2)), int(m.group(1)))
            return d, d
        except ValueError:
            return None, None
    m = re.match(r"^(\d{1,2})\.-(\d{1,2})\.(\d{1,2})$", s)
    if m:
        try:
            return (datetime(year, int(m.group(3)), int(m.group(1))),
                    datetime(year, int(m.group(3)), int(m.group(2))))
        except ValueError:
            return None, None
    m = re.match(r"^(\d{1,2})\.(\d{1,2})\.-(\d{1,2})\.(\d{1,2})$", s)
    if m:
        try:
            return (datetime(year, int(m.group(2)), int(m.group(1))),
                    datetime(year, int(m.group(4)), int(m.group(3))))
        except ValueError:
            return None, None
    return None, None


def ranges_overlap(a0, a1, b0, b1):
    if None in (a0, a1, b0, b1):
        return False
    return a0 <= b1 and b0 <= a1


@st.cache_data
def load_all(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Fahrerwerte"]
    riders = []
    for r in range(4, 50):
        name = ws.cell(r, 1).value
        if not name:
            break
        riders.append({
            "Name": str(name).strip(),
            "EB": ws.cell(r, 2).value, "BE": ws.cell(r, 3).value,
            "MG": ws.cell(r, 4).value, "HÜG": ws.cell(r, 5).value,
            "ZF": ws.cell(r, 6).value, "PRL": ws.cell(r, 7).value,
            "KSP": ws.cell(r, 8).value, "SP": ws.cell(r, 9).value,
            "BES": ws.cell(r, 10).value, "ABF": ws.cell(r, 11).value,
            "ASR": ws.cell(r, 12).value, "AUS": ws.cell(r, 13).value,
            "ZÄH": ws.cell(r, 14).value, "REG": ws.cell(r, 15).value,
            "Fahrerrolle": ws.cell(r, 16).value,
            "Bevorzugte Rennen": ws.cell(r, 17).value,
        })
    df_fw = pd.DataFrame(riders)

    ws = wb["Planung"]
    races = []
    for c in range(3, 60):
        name = ws.cell(3, c).value
        if name is None:
            break
        races.append({
            "name": str(name).strip(),
            "typ": ws.cell(2, c).value,
            "datum": str(ws.cell(4, c).value or ""),
            "etappen": int(ws.cell(5, c).value or 1),
            "col": c,
        })
    for race in races:
        s, e = parse_date_range(race["datum"])
        race["start"], race["end"] = s, e

    plan_rows = []
    for i, rider in enumerate(riders):
        row_idx = 7 + i
        row = {"Fahrer": rider["Name"]}
        for race in races:
            val = ws.cell(row_idx, race["col"]).value
            row[race["name"]] = "X" if val in ("X", "x") else ""
        plan_rows.append(row)
    df_plan = pd.DataFrame(plan_rows)

    ws = wb["Rennprofil"]
    rp = []
    for r in range(2, 60):
        name = ws.cell(r, 1).value
        if not name or str(name).strip() == "":
            continue
        rp.append({
            "Rennen": str(name).strip(),
            "Priorität": ws.cell(r, 3).value,
            "Datum": ws.cell(r, 4).value,
            "Kategorie": ws.cell(r, 5).value,
            "Profil": ws.cell(r, 6).value,
            "Profil-Ampel": ws.cell(r, 7).value,
            "Etappen": ws.cell(r, 8).value,
            "Gesamthärte": ws.cell(r, 10).value,
            "Bergankunft": ws.cell(r, 11).value,
            "Sprintankunft": ws.cell(r, 12).value,
            "Bergetappen": ws.cell(r, 13).value,
            "Mittelgebirge": ws.cell(r, 14).value,
            "Flachetappen": ws.cell(r, 15).value,
            "Kopfsteinpflaster": ws.cell(r, 16).value,
            "Höhenmeter": ws.cell(r, 17).value,
            "Bergfaktor": ws.cell(r, 18).value,
            "Sprintfaktor": ws.cell(r, 19).value,
            "Hügelfaktor": ws.cell(r, 20).value,
            "TT-Faktor": ws.cell(r, 21).value,
            "Windanfälligkeit": ws.cell(r, 22).value,
            "Empfohlene Fahrertypen": ws.cell(r, 23).value,
            "Teilgenommen 2025": ws.cell(r, 24).value,
        })
    df_rp = pd.DataFrame(rp)

    ws = wb["Fahreranalyse"]
    fa = []
    for i, rider in enumerate(riders):
        r = 4 + i
        fa.append({
            "Name": rider["Name"],
            "Fahrerrolle": ws.cell(r, 3).value,
            "Fahrertyp": ws.cell(r, 4).value,
            "Bevorzugte Rennen": ws.cell(r, 5).value,
            "Größe (cm)": ws.cell(r, 7).value,
            "Gewicht (kg)": ws.cell(r, 8).value,
            "Alter": ws.cell(r, 9).value,
            "Nationalität": ws.cell(r, 12).value,
            "Höhen-/Hitzetauglichkeit": ws.cell(r, 13).value,
            "Regenerationsbedarf n. GT": ws.cell(r, 14).value,
            "Formkurve/Saisonziel": ws.cell(r, 15).value,
            "Palmarès": ws.cell(r, 16).value,
        })
    df_fa = pd.DataFrame(fa)

    ws = wb["KI-Regelwerk"]
    lines = []
    for r in range(1, min(80, ws.max_row + 1)):
        for col in (1, 2, 3):
            v = ws.cell(r, col).value
            if v and isinstance(v, str) and len(v.strip()) > 5:
                t = v.strip()
                if t not in lines:
                    lines.append(t)
    regel_text = "\n\n".join(lines[:60]) if lines else "KI-Regelwerk – siehe Excel."

    return {
        "fahrerwerte": df_fw, "rennprofil": df_rp, "planung": df_plan,
        "analyse": df_fa, "races": races, "regelwerk": regel_text,
    }


def _num(v, default=0.0):
    try:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return default
        return float(v)
    except (TypeError, ValueError):
        return default


def compute_eignung(df_fw, df_rp, races):
    """
    Deutliche Spreizung: gewichtete Übereinstimmung, dann Streckung.
    Berglastiges Rennen → BE/MG stark, SP wenig; Sprint → umgekehrt.
    """
    profile = {}
    for _, r in df_rp.iterrows():
        key = str(r["Rennen"]).strip()
        profile[key] = {
            "berg": max(0.5, _num(r.get("Bergfaktor"), 5)),
            "sprint": max(0.5, _num(r.get("Sprintfaktor"), 5)),
            "huegel": max(0.5, _num(r.get("Hügelfaktor"), 5)),
            "tt": max(0.5, _num(r.get("TT-Faktor"), 3)),
            "ksp": 9.0 if str(r.get("Kopfsteinpflaster", "")).lower() in ("ja", "teilweise") else 1.0,
        }

    def factors_for(race_name):
        if race_name in profile:
            return profile[race_name]
        rn = race_name.lower()
        for k, v in profile.items():
            if k.lower() in rn or rn in k.lower():
                return v
        return {"berg": 5, "sprint": 5, "huegel": 5, "tt": 3, "ksp": 1}

    rows = []
    for _, rider in df_fw.iterrows():
        attrs = {
            "berg": _num(rider.get("BE"), 70),
            "sprint": _num(rider.get("SP"), 70),
            "huegel": _num(rider.get("HÜG"), 70),
            "tt": _num(rider.get("ZF"), 70),
            "ksp": _num(rider.get("KSP"), 70),
        }
        # MG fließt leicht in Berg ein
        attrs["berg"] = 0.7 * attrs["berg"] + 0.3 * _num(rider.get("MG"), attrs["berg"])
        row = {"Fahrer": rider["Name"]}
        for race in races:
            f = factors_for(race["name"])
            wsum = sum(f.values()) or 1.0
            # gewichteter Attributwert
            raw = sum(attrs[k] * f[k] for k in f) / wsum
            # Streckung: 55→~20, 70→~55, 85→~90 (deutlichere Unterschiede)
            stretched = (raw - 55.0) * 2.4 + 40.0
            row[race["name"]] = int(max(5, min(98, round(stretched))))
        rows.append(row)
    return pd.DataFrame(rows)


def compute_renntage(plan, races):
    return pd.Series(
        [sum(r["etappen"] for r in races if row.get(r["name"]) == "X") for _, row in plan.iterrows()],
        index=plan.index,
    )


def occupancy(plan, races):
    return {r["name"]: int((plan[r["name"]] == "X").sum()) for r in races}


def find_overlaps(plan, races):
    conflicts = set()
    for idx, row in plan.iterrows():
        assigned = [r for r in races if row.get(r["name"]) == "X"]
        for i in range(len(assigned)):
            for j in range(i + 1, len(assigned)):
                a, b = assigned[i], assigned[j]
                if ranges_overlap(a["start"], a["end"], b["start"], b["end"]):
                    conflicts.add((idx, a["name"]))
                    conflicts.add((idx, b["name"]))
    return conflicts


def style_fahrerwerte(df):
    """Farbskala wie typische PCM/Excel-Bewertung."""
    value_cols = [c for c in df.columns if c not in ("Name", "Fahrerrolle", "Bevorzugte Rennen")]

    def color_val(v):
        try:
            n = float(v)
        except (TypeError, ValueError):
            return ""
        if n >= 85:
            return "background-color:#4ade80;color:#14532d;font-weight:600"
        if n >= 78:
            return "background-color:#86efac;color:#14532d"
        if n >= 72:
            return "background-color:#bbf7d0;color:#166534"
        if n >= 66:
            return "background-color:#fef08a;color:#854d0e"
        if n >= 60:
            return "background-color:#fdba74;color:#9a3412"
        return "background-color:#fecaca;color:#991b1b"

    styler = df.style
    styler = styler.map(color_val, subset=value_cols)
    styler = styler.set_properties(**{
        "background-color": "#f3f4f6",
        "color": "#1f2937",
        "border-color": "#c4c9d1",
    })
    return styler


def build_eign_html(df_eign, races):
    esc = html_lib.escape

    def sc_color(score):
        s = int(score)
        if s >= 80:
            return "#22c55e", "#14532d"
        if s >= 68:
            return "#86efac", "#166534"
        if s >= 55:
            return "#fef08a", "#854d0e"
        if s >= 42:
            return "#fdba74", "#9a3412"
        return "#f87171", "#7f1d1d"

    cells = ['<th style="position:sticky;left:0;z-index:2;background:#e5e7eb;min-width:140px;text-align:left;padding:4px 8px;border:1px solid #c4c9d1;">Fahrer</th>']
    for race in races:
        cells.append(
            f'<th title="{esc(race["name"])}" style="writing-mode:vertical-rl;transform:rotate(180deg);'
            f'height:130px;min-width:30px;background:#e5e7eb;border:1px solid #c4c9d1;font-size:0.75rem;">'
            f'{esc(race["name"][:16])}</th>'
        )
    thead = "<tr>" + "".join(cells) + "</tr>"
    body = []
    for _, row in df_eign.iterrows():
        cells = [
            f'<td style="position:sticky;left:0;z-index:1;background:#eef0f3;font-weight:500;'
            f'text-align:left;white-space:nowrap;padding:3px 8px;border:1px solid #c4c9d1;min-width:140px;">'
            f'{esc(str(row["Fahrer"]))}</td>'
        ]
        for race in races:
            sc = int(row.get(race["name"], 0) or 0)
            bg, fg = sc_color(sc)
            cells.append(
                f'<td style="background:{bg};color:{fg};font-weight:600;text-align:center;'
                f'border:1px solid #c4c9d1;min-width:28px;padding:2px;">{sc}</td>'
            )
        body.append("<tr>" + "".join(cells) + "</tr>")
    return (
        '<div style="overflow-x:auto;border:1.5px solid #9ca3af;border-radius:10px;background:#f3f4f6;">'
        f'<table style="border-collapse:collapse;font-size:0.75rem;color:#1f2937;">{thead}{"".join(body)}</table></div>'
    )


def make_plan_grid_df(plan, races, renntage, occ):
    """DataFrame für AgGrid: Fahrer, Renntage, dann Rennen mit Mehrzeilen-Header-Info in col name."""
    df = plan.copy()
    df.insert(1, "Renntage", renntage.values)
    # kurze Header mit Meta in tooltip via col name
    rename = {}
    for race in races:
        cnt = occ[race["name"]]
        flag = "!" if cnt > MAX_RIDERS else ("=" if cnt == MAX_RIDERS else "")
        short = f"{race['name'][:14]}|{race['datum']}|{race['etappen']}Et|{cnt}/{MAX_RIDERS}{flag}"
        rename[race["name"]] = short
        df[short] = df[race["name"]].fillna("").astype(str).replace({"X": "X", "x": "X"})
        df.drop(columns=[race["name"]], inplace=True)
    return df, rename


# ── Session ────────────────────────────────────────────
if "data" not in st.session_state:
    st.session_state.data = load_all(str(EXCEL))
    st.session_state.plan = st.session_state.data["planung"].copy()
    st.session_state.ai = "Grok (xAI)"
    st.session_state.msg = None
    st.session_state.show_import = False

data = st.session_state.data
races = data["races"]
plan = st.session_state.plan
occ = occupancy(plan, races)
conflicts = find_overlaps(plan, races)
renntage = compute_renntage(plan, races)
df_eign = compute_eignung(data["fahrerwerte"], data["rennprofil"], races)

# ── TOP BAR ────────────────────────────────────────────
c_logo, c_search, c_ki, c_imp, c_exp, c_run, c_form = st.columns([2.0, 1.4, 1.2, 1.0, 1.0, 1.4, 1.4])
with c_logo:
    if LOGO.exists():
        st.image(str(LOGO), width=280)
    else:
        st.markdown("### SIBOS")
with c_search:
    st.text_input("Suche", placeholder="Suchen…", label_visibility="collapsed", key="search")
with c_ki:
    st.session_state.ai = st.selectbox(
        "KI",
        ["Grok (xAI)", "Claude (Anthropic)", "GPT-4o (OpenAI)", "Gemini (Google)", "Lokal (Ollama)"],
        label_visibility="collapsed",
    )
with c_imp:
    if st.button("Importieren", use_container_width=True):
        st.session_state.show_import = not st.session_state.show_import
with c_exp:
    if st.button("Exportieren", use_container_width=True):
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as w:
            plan.to_excel(w, sheet_name="Planung", index=False)
            data["fahrerwerte"].to_excel(w, sheet_name="Fahrerwerte", index=False)
            data["rennprofil"].to_excel(w, sheet_name="Rennprofil", index=False)
            data["analyse"].to_excel(w, sheet_name="Fahreranalyse", index=False)
            df_eign.to_excel(w, sheet_name="Fahrerbewertung", index=False)
        st.download_button(
            "Download Excel", data=buf.getvalue(),
            file_name="SIBOS_Saisonplanung_2026.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
with c_run:
    if st.button("KI-Planung starten", type="primary", use_container_width=True):
        st.session_state.msg = (
            f"**KI-Planung** mit {st.session_state.ai}: Regelwerk Teil A "
            "(X löschen → neu setzen → QC). Automatik folgt; manuell X in der Tabelle setzen."
        )
        st.toast("KI-Planung vorbereitet", icon="⚡")
with c_form:
    if st.button("Formkurven berechnen", type="secondary", use_container_width=True):
        st.session_state.msg = "**Formkurven:** Teil B nach Saisonplanung – Umsetzung folgt."
        st.toast("Formkurven vorbereitet", icon="📈")

# ── IMPORT (inkl. Nachbearbeitung: Farben, Analyse, Eignung neu) ──
if st.session_state.show_import:
    st.markdown("---")
    st.markdown("### Daten importieren")
    ic1, ic2 = st.columns(2)
    with ic1:
        st.markdown("**Excel-Datei**")
        up_x = st.file_uploader("Excel wählen", type=["xlsx"], key="up_xlsx", label_visibility="collapsed")
        if up_x is not None:
            tmp = BASE / "_upload.xlsx"
            tmp.write_bytes(up_x.getbuffer())
            st.session_state.data = load_all(str(tmp))
            st.session_state.plan = st.session_state.data["planung"].copy()
            # Nach Import: Fahrerwerte, Analyse, Eignung werden aus den neuen Daten neu geladen/berechnet
            st.session_state.msg = (
                "Excel importiert. Fahrerwerte (mit Farbskala), Rennprofil, Fahreranalyse und "
                "Eignungsmatrix wurden neu aufgebaut."
            )
            st.session_state.show_import = False
            st.rerun()
    with ic2:
        st.markdown("**Foto / Screenshot**")
        up_img = st.file_uploader(
            "Bild wählen", type=["png", "jpg", "jpeg", "webp"], key="up_img", label_visibility="collapsed"
        )
        if up_img is not None:
            st.session_state.msg = (
                f"Foto „{up_img.name}“ empfangen. Geplant: OCR → Fahrerwerte + Farben + "
                "Fahreranalyse/Eignung automatisch aktualisieren."
            )
            st.session_state.show_import = False
    st.markdown(
        '<div class="import-hint"><b>Anleitung Foto / Screenshot</b><br>'
        "PCM-Bildschirm mit <b>Fahrernamen + Werte</b> (EB, BE, MG, HÜG, ZF, PRL, KSP, SP, …) fotografieren. "
        "Nach dem Einlesen: automatische Farben in Fahrerwerte, Einträge in Analyse/Eignung.</div>",
        unsafe_allow_html=True,
    )

if st.session_state.msg:
    st.info(st.session_state.msg)

# ── PLANUNG (editierbar, Fahrer fixiert) ────────────────
st.markdown("---")
st.markdown("### PCM Season Planner")
st.caption(
    "Zellen anklicken und **X** eintragen oder löschen · "
    "Fahrer/Renntage links bleiben beim horizontalen Scrollen stehen · "
    "KI-Planung startet die automatische Vergabe (in Vorbereitung)"
)

# Header-Hilfe: Belegung
badge_bits = []
for race in races:
    cnt = occ[race["name"]]
    col = "#166534" if cnt == MAX_RIDERS else ("#991b1b" if cnt > MAX_RIDERS else "#374151")
    badge_bits.append(f'<span style="color:{col};margin-right:10px;font-size:0.75rem;">{race["name"][:12]} {cnt}/{MAX_RIDERS}</span>')
st.markdown(
    '<div style="overflow-x:auto;white-space:nowrap;padding:4px 0;">' + "".join(badge_bits) + "</div>",
    unsafe_allow_html=True,
)

# Editierbares Grid
edit_df = plan.copy()
edit_df.insert(1, "Renntage", renntage.values)
# leere Zellen als "" für Klarheit
for r in races:
    edit_df[r["name"]] = edit_df[r["name"]].fillna("").astype(str).replace({"nan": ""})

gb = GridOptionsBuilder.from_dataframe(edit_df)
gb.configure_default_column(editable=True, resizable=True, minWidth=72, cellStyle={"backgroundColor": "#f3f4f6"})
gb.configure_column("Fahrer", editable=False, pinned="left", minWidth=150, lockPinned=True,
                    cellStyle={"backgroundColor": "#e5e7eb", "fontWeight": "600"})
gb.configure_column("Renntage", editable=False, pinned="left", minWidth=80, lockPinned=True,
                    cellStyle={"backgroundColor": "#eef0f3", "fontWeight": "600"})
# Renn-Spalten: nur X erlauben optisch
x_style = JsCode("""
function(params) {
  const v = (params.value || '').toString().toUpperCase();
  if (v === 'X') {
    return {backgroundColor:'#bbf7d0', color:'#14532d', fontWeight:'700', textAlign:'center'};
  }
  return {backgroundColor:'#f3f4f6', textAlign:'center'};
}
""")
for race in races:
    gb.configure_column(
        race["name"],
        editable=True,
        minWidth=78,
        headerTooltip=f"{race['name']} | {race['datum']} | {race['etappen']} Et. | {occ[race['name']]}/{MAX_RIDERS}",
        cellStyle=x_style,
    )
gb.configure_grid_options(
    domLayout="normal",
    ensureDomOrder=True,
    suppressMovableColumns=True,
    rowHeight=28,
    headerHeight=48,
)
grid_opts = gb.build()

grid_resp = AgGrid(
    edit_df,
    gridOptions=grid_opts,
    height=min(50 + 30 * (len(edit_df) + 1), 1100),
    theme="streamlit",
    update_mode=GridUpdateMode.VALUE_CHANGED,
    data_return_mode=DataReturnMode.FILTERED_AND_SORTED,
    fit_columns_on_grid_load=False,
    allow_unsafe_jscode=True,
    key="plan_aggrid",
)

# Änderungen zurückschreiben
try:
    new_df = pd.DataFrame(grid_resp["data"])
    if not new_df.empty and "Fahrer" in new_df.columns:
        changed = False
        new_plan = plan.copy()
        for race in races:
            if race["name"] in new_df.columns:
                series = new_df[race["name"]].apply(
                    lambda v: "X" if str(v).strip().upper() in ("X", "×", "1", "JA", "TRUE") else ""
                )
                if list(series) != list(new_plan[race["name"]].fillna("").astype(str).replace({"nan": ""})):
                    new_plan[race["name"]] = series.values
                    changed = True
        if changed:
            st.session_state.plan = new_plan
            st.rerun()
except Exception:
    pass

st.markdown(
    '<div class="plan-footer">'
    "Grünes X = gesetzt · Belegung in der Kopfzeile (Tooltip) · "
    "Rote Markierung bei Überschneidung folgt nach Speichern · Renntage werden neu berechnet"
    "</div>",
    unsafe_allow_html=True,
)

if conflicts:
    names = sorted({plan.loc[i, "Fahrer"] for i, _ in conflicts})
    st.warning("Zeitliche Überschneidung bei: " + ", ".join(names))

# ── TABS ───────────────────────────────────────────────
st.markdown("---")
st.markdown("### Weitere Tabellenblätter")
t1, t2, t3, t4, t5, t6 = st.tabs(
    ["Rennprofil", "Fahrerwerte", "Fahreranalyse", "Eignung", "Formkurven", "Regelwerk"]
)
TABLE_H = 1100

with t1:
    st.caption("Rennprofil – Fahrerspalte analog fixiert über horizontales Scrollen der Ansicht")
    st.dataframe(data["rennprofil"], use_container_width=True, height=TABLE_H, hide_index=True)

with t2:
    st.caption("Fahrerwerte – Farbskala: hoch = grün, niedrig = rot (wie Excel-Bewertung)")
    st.dataframe(style_fahrerwerte(data["fahrerwerte"]), use_container_width=True, height=TABLE_H)

with t3:
    st.caption("Fahreranalyse")
    st.dataframe(data["analyse"], use_container_width=True, height=TABLE_H, hide_index=True)

with t4:
    st.caption(
        "Eignung 0–100: Fahrerwerte × Rennprofil (Berg/Sprint/Hügel/TT/Pflaster), gestreckt für klare Unterschiede · "
        "Fahrerspalte fixiert beim Scrollen"
    )
    st.markdown(build_eign_html(df_eign, races), unsafe_allow_html=True)

with t5:
    st.caption("Formkurven – Button „Formkurven berechnen“ (Teil B Regelwerk)")
    st.info("Formkurven werden nach der Saisonplanung berechnet.")

with t6:
    st.caption("KI-Regelwerk – Auszug")
    st.text_area("Regelwerk", value=data["regelwerk"], height=500, label_visibility="collapsed")

st.markdown("---")
st.caption("PCM Season Planner 2026 · SIBOS · Import aktualisiert Fahrerwerte/Farben/Analyse/Eignung")
